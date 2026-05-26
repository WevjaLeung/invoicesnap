"""
InvoiceSnap — Backend
FastAPI app: PDF upload, AI extraction, CSV/Excel export, Stripe checkout.
"""

import io
import os
import uuid
import json
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, File, UploadFile, HTTPException, Request, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from extractor import extract_invoice_data
from pricing import create_checkout_session, verify_usage, increment_usage, FREE_LIMIT

app = FastAPI(title="InvoiceSnap", version="0.1.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory usage tracking (privacy-first: nothing persisted to disk)
# In production, use Redis with TTL or signed cookies
usage_store: dict[str, int] = {}

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class ExtractionResult(BaseModel):
    vendor_name: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None
    total_amount: Optional[float] = None
    currency: Optional[str] = "USD"
    tax_amount: Optional[float] = None
    line_items: list[dict] = []
    raw_text_preview: Optional[str] = None
    remaining_free: int = FREE_LIMIT

class ExportRequest(BaseModel):
    data: ExtractionResult
    format: str = "csv"  # csv | xlsx


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------
@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "0.1.0"}


# ---------------------------------------------------------------------------
# Usage tracking
# ---------------------------------------------------------------------------
@app.get("/api/usage")
async def get_usage(request: Request):
    """Return remaining free extractions for this session."""
    session_id = request.headers.get("X-Session-Id", str(uuid.uuid4()))
    used = usage_store.get(session_id, 0)
    return {"used": used, "remaining": max(0, FREE_LIMIT - used), "limit": FREE_LIMIT}


# ---------------------------------------------------------------------------
# Upload & Extract
# ---------------------------------------------------------------------------
@app.post("/api/extract")
async def extract(request: Request, file: UploadFile = File(...)):
    """Upload a PDF invoice and extract structured data via AI."""
    # Validate file type
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are supported.")

    # Session tracking (use header or generate new)
    session_id = request.headers.get("X-Session-Id", str(uuid.uuid4()))
    used = usage_store.get(session_id, 0)

    if used >= FREE_LIMIT:
        return JSONResponse(
            status_code=402,
            content={
                "error": "free_limit_reached",
                "message": f"You've used all {FREE_LIMIT} free extractions. Upgrade to Pro for unlimited.",
                "used": used,
                "limit": FREE_LIMIT,
            },
        )

    # Read file into memory
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:  # 10 MB max
        raise HTTPException(400, "File too large. Max 10 MB.")

    try:
        result = extract_invoice_data(contents, file.filename)
    except Exception as e:
        raise HTTPException(500, f"Extraction failed: {str(e)}")

    # Increment usage
    usage_store[session_id] = used + 1
    result.remaining_free = max(0, FREE_LIMIT - usage_store[session_id])

    return result


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
@app.post("/api/export")
async def export_data(body: ExportRequest):
    """Export extracted invoice data as CSV or Excel."""
    data = body.data

    if body.format == "xlsx":
        return _export_xlsx(data)
    else:
        return _export_csv(data)


def _export_csv(data: ExtractionResult):
    import csv

    output = io.StringIO()
    writer = csv.writer(output)

    # Header info
    writer.writerow(["Field", "Value"])
    writer.writerow(["Vendor", data.vendor_name or ""])
    writer.writerow(["Invoice #", data.invoice_number or ""])
    writer.writerow(["Invoice Date", data.invoice_date or ""])
    writer.writerow(["Due Date", data.due_date or ""])
    writer.writerow(["Currency", data.currency or "USD"])
    writer.writerow(["Tax Amount", data.tax_amount or 0])
    writer.writerow(["Total Amount", data.total_amount or 0])
    writer.writerow([])

    # Line items
    if data.line_items:
        writer.writerow(["Description", "Quantity", "Unit Price", "Amount"])
        for item in data.line_items:
            writer.writerow([
                item.get("description", ""),
                item.get("quantity", ""),
                item.get("unit_price", ""),
                item.get("amount", ""),
            ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=invoice_data.csv"},
    )


def _export_xlsx(data: ExtractionResult):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Header info
    ws["A1"] = "Field"
    ws["B1"] = "Value"
    fields = [
        ("Vendor", data.vendor_name),
        ("Invoice #", data.invoice_number),
        ("Invoice Date", data.invoice_date),
        ("Due Date", data.due_date),
        ("Currency", data.currency),
        ("Tax Amount", data.tax_amount),
        ("Total Amount", data.total_amount),
    ]
    for i, (label, value) in enumerate(fields, 2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    # Line items
    if data.line_items:
        start = len(fields) + 3
        ws[f"A{start}"] = "Description"
        ws[f"B{start}"] = "Quantity"
        ws[f"C{start}"] = "Unit Price"
        ws[f"D{start}"] = "Amount"
        for j, item in enumerate(data.line_items, start + 1):
            ws[f"A{j}"] = item.get("description", "")
            ws[f"B{j}"] = item.get("quantity", "")
            ws[f"C{j}"] = item.get("unit_price", "")
            ws[f"D{j}"] = item.get("amount", "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoice_data.xlsx"},
    )


# ---------------------------------------------------------------------------
# Stripe
# ---------------------------------------------------------------------------
@app.post("/api/create-checkout")
async def checkout():
    """Create a Stripe Checkout session for the Pro plan ($9/mo)."""
    try:
        url = create_checkout_session()
        return {"url": url}
    except Exception as e:
        raise HTTPException(500, f"Stripe error: {str(e)}")


# ---------------------------------------------------------------------------
# Mount frontend at /
# ---------------------------------------------------------------------------
@app.get("/")
async def index():
    from fastapi.responses import FileResponse
    return FileResponse(FRONTEND_DIR / "index.html")


# Static files (CSS, JS, images)
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
